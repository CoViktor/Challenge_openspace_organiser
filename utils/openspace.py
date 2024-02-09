from .table import Table
from .file_utils import call_data
import random as r
import openpyxl


class Openspace():
    def __init__(self, number_of_tables: int= 6, seats_per_table: int= 4, data_source='./new_colleagues.csv') -> None:
        '''Class that creates an environment with tables. 
        Will take number_of_tables as the amount of times a Table class objects should be created and added to a list: self.tables.

        :param number_of_tables: An int that represents the number of tables to be added. Has a default value of 6.
        :param seats_per_table: An int that is taken during the Table object creation to indicate the amount of Seat class objects to be created.
        :param data_source: An str that indicates the pathing to a CSV file where the names to be assigned are listed.
                            This can also take a list of names as input.
        -store(filename): stores the repartition in an excel file. Change source as desired.'''
        self.tables = []
        self.number_of_tables = number_of_tables
        self.seats_per_table = seats_per_table
        self.data_source= data_source
        for x in range(self.number_of_tables):
            table = Table(self.seats_per_table)
            self.tables.append(table)
        self.waiting_line =[]
        
    def __str__(self) -> str:
        '''Method called during a conversion of the object into a chain.'''
        return f'This openspace has {self.number_of_tables} tables, with {self.seats_per_table} seats each.'

    def organize(self, names = None) -> None:
        '''Function that calls the data, shuffles the list and then for each person, 
        goes over the tables and their seats and assigns the person to the first available seat.
        Prints feedback on assignment result, depending on too little or too many seats available.
        
        :param names: A parameter that takes a list of names as an input. If left empty, will default to calling the list
                      from the inputted data source.'''
        if names is None:
            names = call_data(self.data_source)
        colleagues = names
        remaining_seats = 0
        r.shuffle(colleagues)
        for i in range(len(colleagues)):
            if i >= (self.number_of_tables*self.seats_per_table): 
                    self.waiting_line.append(colleagues[i])
            else:
                for n in range(self.number_of_tables):
                    if not self.tables[n].has_free_spot():
                        continue
                    else:
                        self.tables[n].assign_seat(colleagues[i])
                        break
        for table in self.tables:
            remaining_seats += table.left_capacity()
        if len(self.waiting_line) > 0:
               print(f"The following people were unable to take a seat: {self.waiting_line}\nPlease add {len(self.waiting_line)} chairs.")
        elif remaining_seats > 0:
            print(f"Everybody is seated and there are {remaining_seats} empty seats left.")
        else:
             print('Everybody is seated and all seats are filled.')

    def display(self) -> None:
        '''Function that displays each table and how the seats are assigned, as well as any unseated people.'''
        for i in range(len(self.tables)):
            seat_assignments = []
            for seat in self.tables[i].seats:
                seat_assignments.append(seat.occupant)
            print(f"At table {i+1} the seats are assigned as follows:\n{seat_assignments}")
        if len(self.waiting_line) > 0:
            print(f'The following people are unseated: {self.waiting_line}')

    def get_seat_assignments(self) -> None:
        '''Function used for storage for the self.store() function'''
        all_seats=[]
        for i in range(len(self.tables)):
            seat_assignments = [seat.occupant for seat in self.tables[i].seats]
            all_seats.append(["Table " + str(i+1)] + seat_assignments)
        return all_seats

    def store(self, filename: str= 'Seat_assignments.xlsx') -> None:
        '''Function that stores the seat assignments and waiting line in an excel file and prints the pathing.
        :param filename: str that takes as an input the desired output pathing. Insert different storage path if desired.'''
        wb = openpyxl.Workbook()
        ws = wb.active
        all_seats = self.get_seat_assignments()
        for row_index, seats in enumerate(all_seats, start=1):
            for col_index, seat in enumerate(seats, start=1):
                ws.cell(row=row_index, column=col_index, value=seat)
        #Add waiting line
        if len(self.waiting_line) > 0:
            spacer_row = len(all_seats) + 2
            ws.cell(row=spacer_row, column=1, value="Waiting Line:")
            for index, person in enumerate(self.waiting_line, start=spacer_row + 1):
                ws.cell(row=index, column=1, value=person)
        #save the file
        wb.save(filename)
        print(f'Data stored in {filename}')